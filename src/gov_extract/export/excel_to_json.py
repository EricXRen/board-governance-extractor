"""Read a governance Excel workbook and convert it to a BoardGovernanceDocument.

Mirrors the five-sheet structure written by :mod:`gov_extract.export.excel_writer`.
Primarily used to produce ground-truth JSON files from manually prepared spreadsheets.

Fields not represented in the Excel format (``other_positions``, ``date_joined_board``,
``attendance_notes``) are silently set to their defaults (``[]`` / ``None``).

Run directly to convert an Excel workbook to a ground-truth JSON file::

    uv run python src/gov_extract/export/excel_to_json.py report.xlsx \\
        --company "LBG" --year 2025
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import structlog
import typer
from openpyxl import load_workbook
from rich.console import Console

from gov_extract.models.board_summary import BoardSummary
from gov_extract.models.director import (
    AttendanceDetails,
    BiographicalDetails,
    BoardRoleDetails,
    CommitteeAttendance,
    Director,
)
from gov_extract.models.director_election import DirectorElection, DirectorElectionSummary
from gov_extract.models.document import BoardGovernanceDocument, Board
from gov_extract.models.metadata import CompanyMetadata

logger = structlog.get_logger()
_console = Console()

_REQUIRED_SHEETS = [
    "Board Summary",
    "Board Overview",
    "Biographical Details",
    "Committee Memberships",
    "Meeting Attendance",
]

# Sentinel values treated as None / empty
_NULL_STRS = {"N/A", "–", "-", ""}


# ---------------------------------------------------------------------------
# Cell-level parsers
# ---------------------------------------------------------------------------

def _str(val: object) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in _NULL_STRS else None


def _int(val: object) -> int | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in _NULL_STRS:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _float(val: object) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in _NULL_STRS:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _pct(val: object) -> float | None:
    """Parse '88%', '88.0%', or a bare number to a float on the 0–100 scale."""
    if val is None:
        return None
    s = str(val).strip().rstrip("%")
    if s in _NULL_STRS:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _list(val: object) -> list[str]:
    """Split a semicolon-delimited cell into a list, returning [] for empty/N/A."""
    if val is None:
        return []
    s = str(val).strip()
    if s in _NULL_STRS:
        return []
    return [item.strip() for item in s.split(";") if item.strip()]


def _bool_yn(val: object) -> bool | None:
    """Parse 'Yes'/'No' (case-insensitive) to bool, or None for missing values."""
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("yes", "true", "1"):
        return True
    if s in ("no", "false", "0"):
        return False
    return None


def _is_data_row(row: tuple) -> bool:
    """Return True only for genuine data rows (non-empty, non-footer)."""
    val = row[0] if row else None
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and not s.startswith("Source:")


# ---------------------------------------------------------------------------
# Per-sheet readers
# ---------------------------------------------------------------------------

def _read_board_summary(ws: object) -> BoardSummary:
    """Parse the Board Summary sheet into a BoardSummary model."""
    data: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[union-attr]
        if not row or row[0] is None:
            continue
        metric = str(row[0]).strip()
        value = row[1] if len(row) > 1 else None
        if metric and not metric.startswith("Source:"):
            data[metric] = value

    voting_raw = _str(data.get("Voting Standard for Directors"))
    voting = voting_raw if voting_raw in ("Majority", "Plurality") else None

    return BoardSummary(
        ceo_chair_separated=_bool_yn(data.get("CEO and Chair Separated")),
        voting_standard=voting,
        board_size=_int(data.get("Board Size")),
        num_executive_directors=_int(data.get("Executive Directors")),
        num_non_executive_directors=_int(data.get("Non-Executive Directors")),
        num_independent_directors=_int(data.get("Independent Directors")),
        pct_women=_pct(data.get("% Women on Board")),
        pct_independent=_pct(data.get("% Independent Directors")),
        avg_director_age=_float(data.get("Average Director Age")),
        avg_tenure_years=_float(data.get("Average Tenure (years)")),
        notes=_str(data.get("Notes")),
    )


def _read_board_overview(ws: object) -> tuple[dict[str, dict], list[str]]:
    """Parse the Board Overview sheet.

    Returns:
        Tuple of (directors_dict keyed by full_name, ordered list of names).
    """
    directors: dict[str, dict] = {}
    order: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[union-attr]
        if not _is_data_row(row):
            continue
        name = _str(row[0])
        if not name:
            continue

        # Board meetings column stores "attended/scheduled" as a string (col index 12)
        board_mtg = _str(row[12]) if len(row) > 12 else None
        board_attended: int | None = None
        board_scheduled: int | None = None
        if board_mtg and "/" in board_mtg:
            parts = board_mtg.split("/", 1)
            board_attended = _int(parts[0])
            board_scheduled = _int(parts[1])

        directors[name] = {
            "designation": _str(row[1]) if len(row) > 1 else None,
            "board_role_str": _str(row[2]) if len(row) > 2 else None,
            "independence_status": _str(row[3]) if len(row) > 3 else None,
            "year_joined_board": _int(row[4]) if len(row) > 4 else None,
            "tenure_years": _float(row[5]) if len(row) > 5 else None,
            "term_end_year": _int(row[6]) if len(row) > 6 else None,
            "year_end_status": _str(row[7]) if len(row) > 7 else "Active",
            # Fallback committee columns (overridden by Committee Memberships sheet)
            "committees_fallback": _list(row[8]) if len(row) > 8 else [],
            "chairs_fallback": _list(row[9]) if len(row) > 9 else [],
            "num_holding_shares": _int(row[10]) if len(row) > 10 else None,
            "pct_holding_shares": _pct(row[11]) if len(row) > 11 else None,
            "board_meetings_attended": board_attended,
            "board_meetings_scheduled": board_scheduled,
            "board_attendance_pct": _pct(row[13]) if len(row) > 13 else None,
        }
        order.append(name)

    return directors, order


def _read_biographical(ws: object) -> dict[str, dict]:
    """Parse the Biographical Details sheet, keyed by full_name."""
    bios: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[union-attr]
        if not _is_data_row(row):
            continue
        name = _str(row[0])
        if not name:
            continue
        bios[name] = {
            "post_nominals": _str(row[1]) if len(row) > 1 else None,
            "age": _int(row[2]) if len(row) > 2 else None,
            "age_band": _str(row[3]) if len(row) > 3 else None,
            "gender": _str(row[4]) if len(row) > 4 else None,
            "affiliation": _str(row[5]) if len(row) > 5 else None,
            "career_summary": _str(row[6]) if len(row) > 6 else None,
        }
    return bios


def _read_committee_memberships(ws: object) -> dict[str, dict]:
    """Parse the Committee Memberships matrix sheet.

    Returns a dict keyed by full_name with 'committee_memberships' and
    'committee_chair_of' lists (disjoint — a chair entry appears only in
    'committee_chair_of', mirroring the model convention).
    """
    all_rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
    if not all_rows:
        return {}

    header_row = all_rows[0]
    # Committees start at column index 2 (after Name, Designation)
    committees = [str(h).strip() for h in header_row[2:] if h is not None]

    memberships: dict[str, dict] = {}
    for row in all_rows[1:]:
        if not _is_data_row(row):
            continue
        name = _str(row[0])
        if not name:
            continue

        member_of: list[str] = []
        chair_of: list[str] = []
        for i, committee in enumerate(committees):
            col_idx = 2 + i
            cell_val = row[col_idx] if col_idx < len(row) else None
            val = str(cell_val).strip() if cell_val is not None else ""
            if val == "C":
                chair_of.append(committee)
            elif val == "M":
                member_of.append(committee)

        memberships[name] = {
            "committee_memberships": member_of,
            "committee_chair_of": chair_of,
        }
    return memberships


def _read_meeting_attendance(ws: object) -> dict[str, dict]:
    """Parse the Meeting Attendance sheet.

    Committee names are inferred from the repeating header triples
    "{name} Att." / "{name} Sched." / "{name} %".
    ``is_chair`` is left ``False`` here and updated after reading the
    Committee Memberships sheet.
    """
    all_rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
    if not all_rows:
        return {}

    header_row = all_rows[0]
    # Extract committee names from headers like "{name} Att."
    # Fixed columns: Name(0), Designation(1), Board Attended(2),
    #                Board Scheduled(3), Board %(4); then triples at 5, 8, 11, …
    committee_names: list[str] = []
    i = 5
    while i < len(header_row):
        h = header_row[i]
        if h is not None and str(h).strip().endswith(" Att."):
            committee_names.append(str(h).strip()[:-5])
        i += 3

    attendance: dict[str, dict] = {}
    for row in all_rows[1:]:
        if not _is_data_row(row):
            continue
        name = _str(row[0])
        if not name:
            continue

        committee_att: list[dict] = []
        for j, c_name in enumerate(committee_names):
            base = 5 + j * 3
            if base + 2 >= len(row):
                continue
            att_val, sched_val, pct_val = row[base], row[base + 1], row[base + 2]
            if att_val is None or str(att_val).strip() in _NULL_STRS:
                continue
            attended = _int(att_val)
            scheduled = _int(sched_val)
            pct = _pct(pct_val)
            if attended is None and scheduled is None:
                continue
            if pct is None and attended is not None and scheduled:
                pct = round(attended / scheduled * 100.0, 4)
            committee_att.append({
                "committee_name": c_name,
                "meetings_attended": attended or 0,
                "meetings_scheduled": scheduled or 0,
                "attendance_pct": pct or 0.0,
                "is_chair": False,  # back-filled from Committee Memberships
            })

        attendance[name] = {
            "board_meetings_attended": _int(row[2]) if len(row) > 2 else None,
            "board_meetings_scheduled": _int(row[3]) if len(row) > 3 else None,
            "board_attendance_pct": _pct(row[4]) if len(row) > 4 else None,
            "committee_attendance": committee_att,
        }
    return attendance


# ---------------------------------------------------------------------------
# Election sheet readers (optional sheets)
# ---------------------------------------------------------------------------

def _read_election_summary(ws: object) -> DirectorElectionSummary:
    """Parse the Election Summary sheet into a DirectorElectionSummary."""
    data: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[union-attr]
        if not row or row[0] is None:
            continue
        metric = str(row[0]).strip()
        value = row[1] if len(row) > 1 else None
        if metric and not metric.startswith("Source:"):
            data[metric] = value

    return DirectorElectionSummary(
        num_directors_to_elect=_int(data.get("Directors to Elect")),
        candidates_disclosed=_bool_yn(data.get("Candidates Disclosed")),
        incumbent_nominees=_list(data.get("Incumbent Nominees")),
        new_nominees=_list(data.get("New Nominees")),
    )


def _read_election_candidates(ws: object) -> list[Director]:
    """Parse the Election Candidates sheet into a list of Directors.

    The sheet layout is identical to Board Overview, so parsing logic is the
    same as :func:`_read_board_overview` followed by minimal Director construction
    (no biographical or attendance sheets for election candidates).
    """
    candidates: list[Director] = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[union-attr]
        if not _is_data_row(row):
            continue
        name = _str(row[0])
        if not name:
            continue

        board_mtg = _str(row[12]) if len(row) > 12 else None
        board_attended: int | None = None
        board_scheduled: int | None = None
        if board_mtg and "/" in board_mtg:
            parts = board_mtg.split("/", 1)
            board_attended = _int(parts[0])
            board_scheduled = _int(parts[1])

        designation = _str(row[1]) if len(row) > 1 else None
        candidate = Director(
            biographical=BiographicalDetails(full_name=name),
            board_role=BoardRoleDetails(
                designation=designation or "Non-Executive Director",  # type: ignore[arg-type]
                board_role=_str(row[2]) if len(row) > 2 else designation or "Non-Executive Director",
                independence_status=_str(row[3]) if len(row) > 3 else "Independent",  # type: ignore[arg-type]
                year_joined_board=_int(row[4]) if len(row) > 4 else None,
                tenure_years=_float(row[5]) if len(row) > 5 else None,
                term_end_year=_int(row[6]) if len(row) > 6 else None,
                year_end_status=_str(row[7]) if len(row) > 7 else "Active",
                committee_memberships=_list(row[8]) if len(row) > 8 else [],
                committee_chair_of=_list(row[9]) if len(row) > 9 else [],
                other_positions=[],
                num_holding_shares=_int(row[10]) if len(row) > 10 else None,
                pct_holding_shares=_pct(row[11]) if len(row) > 11 else None,
            ),
            attendance=AttendanceDetails(
                board_meetings_attended=board_attended,
                board_meetings_scheduled=board_scheduled,
                board_attendance_pct=_pct(row[13]) if len(row) > 13 else None,
            ),
        )
        candidates.append(candidate)
    return candidates


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def read_excel(
    path: Path,
    company_name: str,
    fiscal_year_end: str,
    filing_type: str = "Annual Report",
    company_ticker: str | None = None,
    report_date: str | None = None,
    source_pdf_path: str = "ground_truth",
) -> BoardGovernanceDocument:
    """Read a governance Excel workbook and return a :class:`BoardGovernanceDocument`.

    The workbook must have the five-sheet structure produced by
    :func:`~gov_extract.export.excel_writer.write_excel` (or a manually prepared
    spreadsheet following the same layout).

    Excel fields not present in the model will default to ``None`` / ``[]``.
    Fields that have no Excel column (``special_roles``, ``date_joined_board``,
    ``attendance_notes``) are set to their model defaults.

    Args:
        path: Path to the ``.xlsx`` file.
        company_name: Company name written to the metadata block.
        fiscal_year_end: ISO-8601 fiscal year end date, e.g. ``"2025-12-31"``.
        filing_type: Filing type label, e.g. ``"Annual Report"``.
        company_ticker: Optional stock ticker symbol.
        report_date: Optional publication date (ISO-8601).
        source_pdf_path: Value written to ``source_pdf_path`` in the metadata.
            Defaults to ``"ground_truth"`` to indicate manual provenance.

    Returns:
        A fully validated :class:`~gov_extract.models.document.BoardGovernanceDocument`.

    Raises:
        ValueError: If a required sheet is missing from the workbook.
    """
    wb = load_workbook(str(path), data_only=True)

    missing = [s for s in _REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Missing required sheets: {missing}")

    board_summary = _read_board_summary(wb["Board Summary"])
    overview_data, director_order = _read_board_overview(wb["Board Overview"])
    bio_data = _read_biographical(wb["Biographical Details"])
    committee_data = _read_committee_memberships(wb["Committee Memberships"])
    attendance_data = _read_meeting_attendance(wb["Meeting Attendance"])

    directors: list[Director] = []
    for name in director_order:
        ov = overview_data[name]
        bio = bio_data.get(name, {})
        cm = committee_data.get(name, {})
        att = attendance_data.get(name, {})

        # Committee Memberships sheet is canonical; fall back to Board Overview columns
        committee_memberships: list[str] = cm.get("committee_memberships") or ov.get("committees_fallback", [])
        committee_chair_of: list[str] = cm.get("committee_chair_of") or ov.get("chairs_fallback", [])

        # Back-fill is_chair on each committee attendance entry
        committee_att_dicts = att.get("committee_attendance", [])
        for ca_dict in committee_att_dicts:
            ca_dict["is_chair"] = ca_dict["committee_name"] in committee_chair_of

        # Attendance figures: Meeting Attendance sheet is canonical over Board Overview
        board_attended = att.get("board_meetings_attended") or ov.get("board_meetings_attended")
        board_scheduled = att.get("board_meetings_scheduled") or ov.get("board_meetings_scheduled")
        board_pct = att.get("board_attendance_pct") or ov.get("board_attendance_pct")

        director = Director(
            biographical=BiographicalDetails(
                full_name=name,
                post_nominals=bio.get("post_nominals"),
                age=bio.get("age"),
                age_band=bio.get("age_band"),
                gender=bio.get("gender"),
                affiliation=bio.get("affiliation"),
                career_summary=bio.get("career_summary"),
            ),
            board_role=BoardRoleDetails(
                designation=ov.get("designation") or "Non-Executive Director",
                board_role=ov.get("board_role_str") or ov.get("designation") or "Non-Executive Director",
                independence_status=ov.get("independence_status") or "Independent",
                year_joined_board=ov.get("year_joined_board"),
                tenure_years=ov.get("tenure_years"),
                term_end_year=ov.get("term_end_year"),
                year_end_status=ov.get("year_end_status") or "Active",
                committee_memberships=committee_memberships,
                committee_chair_of=committee_chair_of,
                other_positions=[],
                num_holding_shares=ov.get("num_holding_shares"),
                pct_holding_shares=ov.get("pct_holding_shares"),
            ),
            attendance=AttendanceDetails(
                board_meetings_attended=board_attended,
                board_meetings_scheduled=board_scheduled,
                board_attendance_pct=board_pct,
                committee_attendance=[CommitteeAttendance(**ca) for ca in committee_att_dicts],
            ),
        )
        directors.append(director)

    company = CompanyMetadata(
        company_name=company_name,
        company_ticker=company_ticker,
        filing_type=filing_type,
        fiscal_year_end=fiscal_year_end,
        report_date=report_date,
        source_pdf_path=source_pdf_path,
        extraction_timestamp=datetime.now(timezone.utc).isoformat(),
        llm_provider="ground_truth",
        llm_model="ground_truth",
    )

    # Optional election sheets
    director_election: DirectorElection | None = None
    if "Election Summary" in wb.sheetnames and "Election Candidates" in wb.sheetnames:
        election_summary = _read_election_summary(wb["Election Summary"])
        election_candidates = _read_election_candidates(wb["Election Candidates"])
        director_election = DirectorElection(
            summary=election_summary,
            candidates=election_candidates,
        )

    doc = BoardGovernanceDocument(
        company=company,
        current_board=Board(summary=board_summary, directors=directors),
        director_election=director_election,
    )
    logger.info(
        "excel_read",
        path=str(path),
        directors=len(directors),
        election_candidates=len(director_election.candidates) if director_election else 0,
    )
    return doc


def from_excel(
    input: str = typer.Argument(
        ...,
        metavar="EXCEL",
        help="Path to a .xlsx file with the five-sheet governance format.",
    ),
    company: str = typer.Option(..., "--company", "-c", help="Company name"),
    year: str = typer.Option(..., "--year", "-y", help="Fiscal year, e.g. 2025"),
    output_dir: str | None = typer.Option(None, "--output-dir", "-o", help="Output directory"),
    eval_id: str | None = typer.Option(
        None,
        "--eval-id",
        help=(
            "Evaluation dataset ID, e.g. 'lbg-fy2025'. "
            "Writes the JSON directly into data/dataset/eval_data/<eval-id>/."
        ),
    ),
    filing_type: str = typer.Option("Annual Report", "--filing-type", help="Filing type"),
    fiscal_year_end: str | None = typer.Option(
        None, "--fiscal-year-end", help="ISO-8601 fiscal year end date"
    ),
    ticker: str | None = typer.Option(None, "--ticker", help="Company ticker symbol"),
    report_date: str | None = typer.Option(None, "--report-date", help="Report date (ISO-8601)"),
    source_pdf: str = typer.Option(
        "ground_truth",
        "--source-pdf",
        help="Value written to source_pdf_path in the metadata block.",
    ),
    config_file: str | None = typer.Option(None, "--config", help="Path to config.yaml"),
) -> None:
    """Convert a governance Excel workbook to a ground-truth JSON file.

    Reads a .xlsx file that has the five-sheet structure produced by the extract
    command and writes a validated BoardGovernanceDocument JSON.  Intended for
    curating ground-truth files for the evaluation dataset.

    Examples:\n
        uv run python excel_to_json.py report.xlsx --company "LBG" --year 2025\n
        uv run python excel_to_json.py report.xlsx --company "LBG" --year 2025 --eval-id lbg-fy2025
    """
    from gov_extract.config import get_config
    from gov_extract.export.json_writer import output_path as json_path
    from gov_extract.export.json_writer import write_json

    cfg = get_config(Path(config_file) if config_file else None)

    xlsx_path = Path(input)
    if not xlsx_path.exists():
        _console.print(f"[bold red]Error:[/bold red] File not found: {xlsx_path}")
        raise typer.Exit(1)

    resolved_fye = fiscal_year_end or f"{year}-12-31"

    _console.print(
        f"[bold cyan]from-excel[/bold cyan] "
        f"[green]{company}[/green] ({year})  ←  {xlsx_path.name}"
    )

    with _console.status("Reading Excel workbook..."):
        doc = read_excel(
            path=xlsx_path,
            company_name=company,
            fiscal_year_end=resolved_fye,
            filing_type=filing_type,
            company_ticker=ticker,
            report_date=report_date,
            source_pdf_path=source_pdf,
        )

    _console.print(f"  Directors read: [bold green]{len(doc.current_board.directors)}[/bold green]")

    if eval_id:
        dest_dir = Path(cfg.output.eval_dataset_dir) / eval_id
    else:
        dest_dir = Path(output_dir) if output_dir else Path(cfg.output.default_dir)

    dest_dir.mkdir(parents=True, exist_ok=True)
    json_out = json_path(company, year, dest_dir)

    with _console.status("Writing JSON..."):
        write_json(doc, json_out)

    _console.print("\n[bold green]Done![/bold green]")
    _console.print(f"  JSON: {json_out}")
    if eval_id:
        _console.print(f"  (saved to eval dataset: {dest_dir})")


if __name__ == "__main__":
    typer.run(from_excel)

"""Write a four-sheet Excel workbook matching the LBG reference format."""

from __future__ import annotations

from pathlib import Path

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from gov_extract.models.board_summary import BoardSummary
from gov_extract.models.director import Director
from gov_extract.models.director_election import DirectorElection
from gov_extract.models.document import BoardGovernanceDocument

logger = structlog.get_logger()

# Formatting constants matching the reference file
HDR_BG = "1B3A6B"
HDR_FG = "FFFFFF"
EXEC_BG = "FFF3CD"
CHAIR_BG = "E8EAF6"
ALT_BG = "F2F7FC"
ATT_GREEN = "C8E6C9"
ATT_YELLOW = "FFF9C4"
ATT_RED = "FFCDD2"
FONT_NAME = "Arial"


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HDR_BG)


def _row_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _att_fill(pct: float | None) -> PatternFill:
    if pct is None:
        return PatternFill()
    if pct >= 100.0:
        return _row_fill(ATT_GREEN)
    if pct >= 80.0:
        return _row_fill(ATT_YELLOW)
    return _row_fill(ATT_RED)


def _hdr_font() -> Font:
    return Font(name=FONT_NAME, bold=True, color=HDR_FG, size=10)


def _cell_font(bold: bool = False) -> Font:
    return Font(name=FONT_NAME, bold=bold, size=10)


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _director_fill(director: Director) -> PatternFill:
    designation = director.board_role.designation
    if designation == "Chair":
        return _row_fill(CHAIR_BG)
    if designation == "Executive Director":
        return _row_fill(EXEC_BG)
    return PatternFill()  # NED — white, alternating handled in caller


def _write_header(ws: object, headers: list[str]) -> None:
    """Write a styled header row to a worksheet."""

    ws = ws  # type: ignore[assignment]
    ws.append(headers)  # type: ignore[union-attr]
    for cell in ws[1]:  # type: ignore[index]
        cell.fill = _header_fill()
        cell.font = _hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 30  # type: ignore[union-attr]


def _apply_row_style(ws: object, row_idx: int, fill: PatternFill, alt: bool = False) -> None:

    if (not fill.fgColor or fill.fgColor.value == "00000000") and alt:
        fill = _row_fill(ALT_BG)
    for cell in ws[row_idx]:  # type: ignore[index]
        cell.fill = fill
        cell.font = _cell_font()
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autofit_columns(ws: object, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:  # type: ignore[union-attr]
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))  # type: ignore[union-attr]


def _format_pct(pct: float | None) -> str:
    if pct is None:
        return "N/A"
    return f"{pct:.0f}%"


def _footer_text(doc: BoardGovernanceDocument) -> str:
    ts = doc.company.extraction_timestamp[:10]
    return (
        f"Source: {doc.company.source_pdf_path}  |  "
        f"Extracted: {ts}  |  "
        f"Provider: {doc.company.llm_provider} / {doc.company.llm_model}"
    )


def _add_footer(ws: object, row: int, col_count: int, text: str) -> None:
    ws.append([""])  # type: ignore[union-attr]
    ws.append([text])  # type: ignore[union-attr]
    footer_row = row + 2
    cell = ws.cell(row=footer_row, column=1)  # type: ignore[union-attr]
    cell.value = text
    cell.font = Font(name=FONT_NAME, italic=True, size=8, color="888888")


def _write_board_summary(wb: Workbook, doc: BoardGovernanceDocument) -> None:
    """Write the Board Summary sheet — one row per metric."""
    ws = wb.create_sheet("Board Summary")
    _write_header(ws, ["Metric", "Value"])
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    summary: BoardSummary = doc.board_summary

    def _pct(v: float | None) -> str:
        return f"{v:.1f}%" if v is not None else "N/A"

    def _yn(v: bool | None) -> str:
        if v is None:
            return "N/A"
        return "Yes" if v else "No"

    rows = [
        ("CEO and Chair Separated", _yn(summary.ceo_chair_separated)),
        ("Voting Standard for Directors", summary.voting_standard or "N/A"),
        ("Board Size", summary.board_size if summary.board_size is not None else "N/A"),
        (
            "Executive Directors",
            summary.num_executive_directors if summary.num_executive_directors is not None else "N/A",
        ),
        (
            "Non-Executive Directors",
            summary.num_non_executive_directors if summary.num_non_executive_directors is not None else "N/A",
        ),
        (
            "Independent Directors",
            summary.num_independent_directors if summary.num_independent_directors is not None else "N/A",
        ),
        ("% Women on Board", _pct(summary.pct_women)),
        ("% Independent Directors", _pct(summary.pct_independent)),
        ("Average Director Age", f"{summary.avg_director_age:.1f}" if summary.avg_director_age is not None else "N/A"),
        ("Average Tenure (years)", f"{summary.avg_tenure_years:.1f}" if summary.avg_tenure_years is not None else "N/A"),
        ("Notes", summary.notes or ""),
    ]

    for i, (metric, value) in enumerate(rows, 2):
        ws.cell(row=i, column=1).value = metric
        ws.cell(row=i, column=2).value = value
        fill = _row_fill(ALT_BG) if i % 2 == 0 else PatternFill()
        for col in (1, 2):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.font = _cell_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _add_footer(ws, len(rows) + 2, 2, _footer_text(doc))


def _write_board_overview(wb: Workbook, doc: BoardGovernanceDocument) -> None:
    ws = wb.create_sheet("Board Overview")
    headers = [
        "Name",
        "Designation",
        "Board Role",
        "Independence",
        "Year Joined",
        "Tenure (yrs)",
        "Term End Year",
        "Status",
        "Committees",
        "Chairs",
        "Shares Held",
        "% Shares",
        "Board Meetings",
        "Attendance %",
    ]
    _write_header(ws, headers)

    for i, director in enumerate(doc.directors, 1):
        bio = director.biographical
        role = director.board_role
        att = director.attendance

        committees = "; ".join(role.committee_memberships)
        chairs = "; ".join(role.committee_chair_of)
        board_mtg = ""
        if att.board_meetings_attended is not None and att.board_meetings_scheduled is not None:
            board_mtg = f"{att.board_meetings_attended}/{att.board_meetings_scheduled}"

        row = [
            bio.full_name,
            role.designation,
            role.board_role,
            role.independence_status,
            role.year_joined_board,
            role.tenure_years,
            role.term_end_year,
            role.year_end_status,
            committees,
            chairs,
            role.num_holding_shares,
            _format_pct(role.pct_holding_shares) if role.pct_holding_shares is not None else "N/A",
            board_mtg,
            _format_pct(att.board_attendance_pct),
        ]
        ws.append(row)
        row_idx = i + 1
        fill = _director_fill(director)
        _apply_row_style(ws, row_idx, fill, alt=(i % 2 == 0))

        # Traffic-light attendance cell
        att_cell = ws.cell(row=row_idx, column=14)
        att_cell.fill = _att_fill(att.board_attendance_pct)

    _autofit_columns(ws)
    _add_footer(ws, len(doc.directors) + 1, len(headers), _footer_text(doc))


def _write_biographical(wb: Workbook, doc: BoardGovernanceDocument) -> None:
    ws = wb.create_sheet("Biographical Details")
    headers = [
        "Name",
        "Post-Nominals",
        "Age",
        "Age Band",
        "Gender",
        "Affiliation",
        "Career Summary",
    ]
    _write_header(ws, headers)

    for i, director in enumerate(doc.directors, 1):
        bio = director.biographical
        row = [
            bio.full_name,
            bio.post_nominals,
            bio.age,
            bio.age_band,
            bio.gender,
            bio.affiliation,
            bio.career_summary,
        ]
        ws.append(row)
        fill = _director_fill(director)
        _apply_row_style(ws, i + 1, fill, alt=(i % 2 == 0))

    _autofit_columns(ws, max_width=60)
    _add_footer(ws, len(doc.directors) + 1, len(headers), _footer_text(doc))


def _write_committee_memberships(wb: Workbook, doc: BoardGovernanceDocument) -> None:
    ws = wb.create_sheet("Committee Memberships")

    # Collect all unique committee names
    all_committees: list[str] = []
    for director in doc.directors:
        for c in director.board_role.committee_memberships + director.board_role.committee_chair_of:
            if c not in all_committees:
                all_committees.append(c)
    all_committees = sorted(all_committees)

    headers = ["Name", "Designation"] + all_committees
    _write_header(ws, headers)

    for i, director in enumerate(doc.directors, 1):
        role = director.board_role
        row: list[str] = [director.biographical.full_name, role.designation]
        for committee in all_committees:
            if committee in role.committee_chair_of:
                row.append("C")
            elif committee in role.committee_memberships:
                row.append("M")
            else:
                row.append("–")
        ws.append(row)
        fill = _director_fill(director)
        _apply_row_style(ws, i + 1, fill, alt=(i % 2 == 0))

    _autofit_columns(ws)
    _add_footer(ws, len(doc.directors) + 1, len(headers), _footer_text(doc))


def _write_meeting_attendance(wb: Workbook, doc: BoardGovernanceDocument) -> None:
    ws = wb.create_sheet("Meeting Attendance")

    # Collect all committee names from attendance records
    all_committees: list[str] = []
    for director in doc.directors:
        for ca in director.attendance.committee_attendance:
            if ca.committee_name not in all_committees:
                all_committees.append(ca.committee_name)
    all_committees = sorted(all_committees)

    headers = ["Name", "Designation", "Board Attended", "Board Scheduled", "Board %"]
    for c in all_committees:
        headers += [f"{c} Att.", f"{c} Sched.", f"{c} %"]
    _write_header(ws, headers)

    for i, director in enumerate(doc.directors, 1):
        att = director.attendance
        row: list[object] = [
            director.biographical.full_name,
            director.board_role.designation,
            att.board_meetings_attended,
            att.board_meetings_scheduled,
            _format_pct(att.board_attendance_pct),
        ]

        # Build committee lookup
        ca_map = {ca.committee_name: ca for ca in att.committee_attendance}
        for c in all_committees:
            ca = ca_map.get(c)
            if ca:
                row += [ca.meetings_attended, ca.meetings_scheduled, _format_pct(ca.attendance_pct)]
            else:
                row += ["N/A", "N/A", "N/A"]

        ws.append(row)
        row_idx = i + 1
        fill = _director_fill(director)
        _apply_row_style(ws, row_idx, fill, alt=(i % 2 == 0))

        # Traffic-light: board attendance %
        board_pct_col = 5
        ws.cell(row=row_idx, column=board_pct_col).fill = _att_fill(att.board_attendance_pct)

        # Traffic-light: committee attendance %
        for j, c in enumerate(all_committees):
            ca = ca_map.get(c)
            pct_col = 6 + j * 3 + 2  # third column of each triple
            ws.cell(row=row_idx, column=pct_col).fill = _att_fill(ca.attendance_pct if ca else None)

    _autofit_columns(ws)
    _add_footer(ws, len(doc.directors) + 1, len(headers), _footer_text(doc))


def _write_election_summary(wb: Workbook, election: DirectorElection, footer: str) -> None:
    """Write the Election Summary sheet — two-column metric/value table."""
    ws = wb.create_sheet("Election Summary")
    _write_header(ws, ["Metric", "Value"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    summary = election.summary

    def _yn(v: bool | None) -> str:
        if v is None:
            return "N/A"
        return "Yes" if v else "No"

    rows = [
        ("Directors to Elect", summary.num_directors_to_elect if summary.num_directors_to_elect is not None else "N/A"),
        ("Candidates Disclosed", _yn(summary.candidates_disclosed)),
        ("Incumbent Nominees", "; ".join(summary.incumbent_nominees) or "N/A"),
        ("New Nominees", "; ".join(summary.new_nominees) or "N/A"),
        ("Total Candidates", len(election.candidates)),
    ]

    for i, (metric, value) in enumerate(rows, 2):
        ws.cell(row=i, column=1).value = metric
        ws.cell(row=i, column=2).value = value
        fill = _row_fill(ALT_BG) if i % 2 == 0 else PatternFill()
        for col in (1, 2):
            cell = ws.cell(row=i, column=col)
            cell.fill = fill
            cell.font = _cell_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _add_footer(ws, len(rows) + 2, 2, footer)


def _write_election_candidates(wb: Workbook, election: DirectorElection, footer: str) -> None:
    """Write the Election Candidates sheet — same layout as Board Overview."""
    ws = wb.create_sheet("Election Candidates")
    headers = [
        "Name",
        "Designation",
        "Board Role",
        "Independence",
        "Year Joined",
        "Tenure (yrs)",
        "Term End Year",
        "Status",
        "Committees",
        "Chairs",
        "Shares Held",
        "% Shares",
        "Board Meetings",
        "Attendance %",
    ]
    _write_header(ws, headers)

    for i, candidate in enumerate(election.candidates, 1):
        bio = candidate.biographical
        role = candidate.board_role
        att = candidate.attendance

        committees = "; ".join(role.committee_memberships)
        chairs = "; ".join(role.committee_chair_of)
        board_mtg = ""
        if att.board_meetings_attended is not None and att.board_meetings_scheduled is not None:
            board_mtg = f"{att.board_meetings_attended}/{att.board_meetings_scheduled}"

        row = [
            bio.full_name,
            role.designation,
            role.board_role,
            role.independence_status,
            role.year_joined_board,
            role.tenure_years,
            role.term_end_year,
            role.year_end_status,
            committees,
            chairs,
            role.num_holding_shares,
            _format_pct(role.pct_holding_shares) if role.pct_holding_shares is not None else "N/A",
            board_mtg,
            _format_pct(att.board_attendance_pct),
        ]
        ws.append(row)
        row_idx = i + 1
        fill = _director_fill(candidate)
        _apply_row_style(ws, row_idx, fill, alt=(i % 2 == 0))

        att_cell = ws.cell(row=row_idx, column=14)
        att_cell.fill = _att_fill(att.board_attendance_pct)

    _autofit_columns(ws)
    _add_footer(ws, len(election.candidates) + 1, len(headers), footer)


def write_excel(doc: BoardGovernanceDocument, path: Path) -> Path:
    """Write a four-sheet Excel workbook for a BoardGovernanceDocument.

    Args:
        doc: The governance document to export.
        path: Output .xlsx file path.

    Returns:
        The path written to.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_board_summary(wb, doc)
    _write_board_overview(wb, doc)
    _write_biographical(wb, doc)
    _write_committee_memberships(wb, doc)
    _write_meeting_attendance(wb, doc)

    if doc.director_election is not None:
        footer = _footer_text(doc)
        _write_election_summary(wb, doc.director_election, footer)
        _write_election_candidates(wb, doc.director_election, footer)

    wb.save(str(path))
    logger.info("excel_written", path=str(path), directors=len(doc.directors))
    return path


def output_path(company_name: str, fiscal_year: str, output_dir: Path) -> Path:
    """Build the canonical output file path for an Excel export.

    Args:
        company_name: Company name.
        fiscal_year: e.g. "2025".
        output_dir: Directory for output files.

    Returns:
        Path like output_dir/CompanyName_2025_Board_Governance.xlsx.
    """
    safe_name = company_name.replace(" ", "")
    return output_dir / f"{safe_name}_{fiscal_year}_Board_Governance.xlsx"



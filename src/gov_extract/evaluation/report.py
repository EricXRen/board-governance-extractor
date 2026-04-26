"""Write evaluation reports: JSON, Excel, and rich stdout summary."""

from __future__ import annotations

import dataclasses
import json
from pathlib import Path
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rich.console import Console
from rich.table import Table

from gov_extract.evaluation.evaluator import CorpusResult, DocumentResult

logger = structlog.get_logger()

# Colours for the evaluation report Excel
FAIL_FILL = PatternFill("solid", fgColor="FFCDD2")  # below_threshold
HALL_FILL = PatternFill("solid", fgColor="FFE0B2")  # hallucination
FN_FILL = PatternFill("solid", fgColor="FFF9C4")  # false_negative
PASS_FILL = PatternFill("solid", fgColor="C8E6C9")  # pass
HDR_FILL = PatternFill("solid", fgColor="1B3A6B")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="Arial", size=10)


def _dataclass_to_dict(obj: Any) -> Any:
    """Recursively convert dataclasses (and nested structures) to dicts."""
    if dataclasses.is_dataclass(obj) and not isinstance(obj, type):
        return {k: _dataclass_to_dict(v) for k, v in dataclasses.asdict(obj).items()}
    if isinstance(obj, list):
        return [_dataclass_to_dict(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _dataclass_to_dict(v) for k, v in obj.items()}
    return obj


def write_json_report(result: DocumentResult | CorpusResult, output_dir: Path) -> Path:
    """Write the evaluation result to a JSON file.

    Args:
        result: DocumentResult or CorpusResult.
        output_dir: Directory to write the report into.

    Returns:
        Path to the written file.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "evaluation_report.json"
    data = _dataclass_to_dict(result)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    logger.info("eval_report_json_written", path=str(path))
    return path


def write_excel_report(result: DocumentResult, output_dir: Path) -> Path:
    """Write a tabular evaluation report to an Excel file.

    Args:
        result: DocumentResult with all field-level results.
        output_dir: Directory to write the report into.

    Returns:
        Path to the written file.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "evaluation_report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Field Results"  # type: ignore[union-attr]

    headers = [
        "Director",
        "Field Path",
        "Metric",
        "Predicted",
        "Ground Truth",
        "Score",
        "Passed",
        "Failure Mode",
    ]
    ws.append(headers)  # type: ignore[union-attr]
    for cell in ws[1]:  # type: ignore[index]
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for dr in result.director_results:
        for fr in dr.field_results:
            row = [
                dr.director_name,
                fr.field_path,
                fr.metric_used,
                str(fr.predicted_value) if fr.predicted_value is not None else "",
                str(fr.ground_truth_value) if fr.ground_truth_value is not None else "",
                round(fr.score, 4),
                "PASS" if fr.passed else "FAIL",
                fr.failure_mode or "",
            ]
            ws.append(row)  # type: ignore[union-attr]
            row_idx = ws.max_row  # type: ignore[union-attr]

            fill = PASS_FILL
            if fr.failure_mode == "below_threshold":
                fill = FAIL_FILL
            elif fr.failure_mode == "hallucination":
                fill = HALL_FILL
            elif fr.failure_mode == "false_negative":
                fill = FN_FILL

            for cell in ws[row_idx]:  # type: ignore[index]
                cell.fill = fill
                cell.font = CELL_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = [25, 40, 20, 30, 30, 8, 8, 18]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(i)].width = w  # type: ignore[union-attr]

    wb.save(str(path))
    logger.info("eval_report_excel_written", path=str(path))
    return path


def print_rich_summary(result: DocumentResult) -> None:
    """Print a rich formatted summary to stdout.

    Args:
        result: DocumentResult to summarise.
    """
    console = Console()

    # Panel 1: Headline metrics
    headline = Table(title=f"[bold]Evaluation: {result.company_name}[/bold]", show_header=True)
    headline.add_column("Metric", style="cyan")
    headline.add_column("Value", style="green")

    headline.add_row("Document Field Pass Rate", f"{result.document_field_pass_rate:.1%}")
    headline.add_row("Document Perfect Match", str(result.document_perfect_match))
    headline.add_row("Director Perfect Match Rate", f"{result.director_perfect_match_rate:.1%}")
    headline.add_row("False Negative Rate", f"{result.false_negative_rate:.1%}")
    headline.add_row("Hallucination Rate", f"{result.hallucination_rate:.1%}")
    headline.add_row("Directors Evaluated", str(len(result.director_results)))
    console.print(headline)

    # Panel 2: Per-field-type pass rates (worst first)
    type_table = Table(title="Per-Field-Type Pass Rates (worst first)", show_header=True)
    type_table.add_column("Category", style="cyan")
    type_table.add_column("Pass Rate", style="green")

    for cat, rate in sorted(result.per_field_type_pass_rate.items(), key=lambda x: x[1]):
        colour = "green" if rate >= 0.90 else ("yellow" if rate >= 0.75 else "red")
        type_table.add_row(cat, f"[{colour}]{rate:.1%}[/{colour}]")
    console.print(type_table)

    # Panel 3: Five worst individual fields
    worst_fields = sorted(result.per_field_pass_rate.items(), key=lambda x: x[1])[:5]
    worst_table = Table(title="Five Worst-Performing Fields", show_header=True)
    worst_table.add_column("Field Path", style="cyan")
    worst_table.add_column("Pass Rate", style="red")

    for fp, rate in worst_fields:
        worst_table.add_row(fp, f"{rate:.1%}")
    console.print(worst_table)


def write_evaluation_report(result: DocumentResult | CorpusResult, output_dir: Path) -> None:
    """Write all three evaluation report artefacts.

    Args:
        result: Evaluation result (document or corpus level).
        output_dir: Output directory.
    """
    write_json_report(result, output_dir)
    if isinstance(result, DocumentResult):
        write_excel_report(result, output_dir)
        print_rich_summary(result)
